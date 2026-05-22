"""
backfill_history.py — Build initial trend_history.json from existing Run 1 & 2 Excel files.

Reads the Detail tab from each tribe Excel to reconstruct per-initiative data,
then calls compute_run_entry() (imported from build_reports logic) to produce
the same schema that build_reports.py writes going forward.

Run once before uploading trend_history.json to Drive.
"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')

from collections import defaultdict
from openpyxl import load_workbook

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
TREND_FILE = os.path.join(BASE_DIR, 'trend_history.json')

FIELD_LABELS = [
    ('Summary',               'summary'),
    ('Parent',                'parent'),
    ('Product Lead',          'customfield_12121'),
    ('Engineering Lead',      'customfield_12122'),
    ('Roadmap Priority',      'customfield_12012'),
    ('Impact (PDT)',          'customfield_12178'),
    ('Country (PDT)',         'customfield_12112'),
    ('Areas Impacted',        'customfield_12110'),
    ('Investment Category',   'customfield_10709'),
    ('Start Date',            'customfield_10025'),
    ('Description / PRD',     'desc_or_prd'),
    ('Tribes Impacted',       'customfield_12109'),
    ('End of Definition Date','customfield_15460'),
    ('Health Status',         'customfield_12111'),
    ('Planned Release Date',  'customfield_12114'),
    ('Short Status Update',   'customfield_14447'),
    ('Due Date',              'duedate'),
]
LABEL_TO_FID = {lbl: fid for lbl, fid in FIELD_LABELS}
FIELD_LABELS_LIST = [lbl for lbl, _ in FIELD_LABELS]

TRIBES_ORDER = ['Player Engagement', 'Transact', 'Fraud Prevention',
                'Retail/Multichannel', 'Manage', 'Unassigned']

TRIBE_SLUGS = {
    'Player Engagement':   'PlayerEngagement',
    'Transact':            'Transact',
    'Fraud Prevention':    'FraudPrevention',
    'Retail/Multichannel': 'RetailMultichannel',
    'Manage':              'Manage',
}

RUNS = [
    {
        'run': 1,
        'date': '2026-04-25',
        'folder': r'C:\Users\JonathanVince\Desktop\Jira Data Quality\PLAYER Data Quality — Run 1 — 2026-04-25',
    },
    {
        'run': 2,
        'date': '2026-05-04',
        'folder': os.path.join(BASE_DIR, 'PLAYER Data Quality — Run 2 — 2026-05-04'),
    },
]


def read_detail_tab(xlsx_path, tribe_name):
    """Read the Detail tab of a tribe Excel and return a list of row dicts."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if 'Detail' not in wb.sheetnames:
        print(f'  WARN: no Detail tab in {os.path.basename(xlsx_path)}')
        wb.close()
        return []

    ws = wb['Detail']
    rows_out = []
    header_row = None

    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri == 0:
            header_row = list(row)
            continue
        if not row or not row[1]:   # Issue Key is col B (index 1)
            continue

        score_raw = row[0]
        score = int(score_raw) if isinstance(score_raw, (int, float)) else 0

        key     = str(row[1]).strip()
        summary = str(row[2]).strip() if row[2] else ''
        status  = str(row[3]).strip() if row[3] else ''
        squad   = str(row[4]).strip() if row[4] else 'Unassigned'
        created = str(row[5]).strip() if row[5] else ''

        # Reconstruct fields dict and missing list from field columns (index 6+)
        fields  = {}
        missing = []
        for ci, (lbl, fid) in enumerate(FIELD_LABELS):
            cell_val = row[6 + ci] if (6 + ci) < len(row) else None
            if cell_val == 'N/A' or cell_val is None:
                pass  # field not applicable for this initiative's tier
            elif str(cell_val).strip() == '✓':
                fields[fid] = True
            else:
                fields[fid] = False
                missing.append(lbl)

        rows_out.append({
            'key': key, 'summary': summary, 'status': status,
            'squad': squad, 'tribe': tribe_name, 'created': created,
            'score': score, 'fields': fields, 'missing': missing,
        })

    wb.close()
    return rows_out


def compute_run_entry(rows, date_str, run_num):
    total = len(rows)
    if not total:
        return None
    with_gaps  = sum(1 for r in rows if r['score'] < 100)
    compliant  = total - with_gaps
    avg_score  = round(sum(r['score'] for r in rows) / total)
    gap_counts = {}
    for lbl, fid in FIELD_LABELS:
        n = sum(1 for r in rows if fid in r.get('fields', {}) and not r['fields'][fid])
        if n:
            gap_counts[lbl] = n

    tribes_out = {}
    for tribe in TRIBES_ORDER:
        t_rows = [r for r in rows if r.get('tribe', 'Unassigned') == tribe]
        if not t_rows:
            continue
        t_total     = len(t_rows)
        t_with_gaps = sum(1 for r in t_rows if r['score'] < 100)
        t_avg       = round(sum(r['score'] for r in t_rows) / t_total)
        t_gap_c     = {}
        for lbl, fid in FIELD_LABELS:
            n = sum(1 for r in t_rows if fid in r.get('fields', {}) and not r['fields'][fid])
            if n:
                t_gap_c[lbl] = n

        squads_out = {}
        squad_map  = defaultdict(list)
        for r in t_rows:
            squad_map[r.get('squad', 'Unassigned')].append(r)
        for squad, s_rows in squad_map.items():
            s_avg   = round(sum(r['score'] for r in s_rows) / len(s_rows))
            s_with  = sum(1 for r in s_rows if r['score'] < 100)
            s_gap_c = {}
            for lbl, fid in FIELD_LABELS:
                n = sum(1 for r in s_rows if fid in r.get('fields', {}) and not r['fields'][fid])
                if n:
                    s_gap_c[lbl] = n
            squads_out[squad] = {'total': len(s_rows), 'avg_score': s_avg,
                                 'with_gaps': s_with, 'gap_counts': s_gap_c}

        tribes_out[tribe] = {
            'total': t_total, 'avg_score': t_avg,
            'with_gaps': t_with_gaps, 'compliant': t_total - t_with_gaps,
            'gap_counts': t_gap_c, 'squads': squads_out,
        }

    active = sorted(
        [r for r in rows if r.get('status', '').lower() != 'done' and r.get('missing')],
        key=lambda r: len(r.get('missing', [])), reverse=True
    )
    top_offenders = [
        {'key': r['key'], 'summary': r['summary'],
         'tribe': r.get('tribe', 'Unassigned'), 'squad': r.get('squad', ''),
         'score': r['score'], 'missing_count': len(r['missing']), 'missing': r['missing']}
        for r in active[:20]
    ]

    return {
        'run': run_num, 'date': date_str,
        'total': total, 'with_gaps': with_gaps, 'compliant': compliant,
        'avg_score': avg_score, 'gap_counts': gap_counts,
        'tribes': tribes_out, 'top_offenders': top_offenders,
        'source': 'backfill',
    }


def main():
    history = {'schema_version': 1, 'runs': []}

    for run_def in RUNS:
        run_num  = run_def['run']
        date_str = run_def['date']
        folder   = run_def['folder']

        if not os.path.isdir(folder):
            print(f'Run {run_num} folder not found: {folder}')
            print('  Skipping — upload the tribe Excels to Drive; Apps Script backfillHistory() will handle it.')
            continue

        print(f'\nRun {run_num} ({date_str}): reading from {folder}')
        all_rows = []

        for tribe, slug in TRIBE_SLUGS.items():
            fname = f'PLAYER_DataQuality_{slug}_{date_str}.xlsx'
            fpath = os.path.join(folder, fname)
            if not os.path.exists(fpath):
                print(f'  SKIP: {fname} not found')
                continue
            rows = read_detail_tab(fpath, tribe)
            all_rows.extend(rows)
            squads = len({r['squad'] for r in rows})
            avg    = round(sum(r['score'] for r in rows) / len(rows)) if rows else 0
            print(f'  {tribe}: {len(rows)} initiatives, {squads} squads, avg {avg}%')

        if not all_rows:
            print(f'  No data — skipping Run {run_num}')
            continue

        entry = compute_run_entry(all_rows, date_str, run_num)
        if entry:
            history['runs'].append(entry)
            print(f'  Entry computed: {entry["total"]} total, avg {entry["avg_score"]}%, '
                  f'{len(entry["top_offenders"])} offenders')

    if not history['runs']:
        print('\nNo runs processed. Exiting without writing trend_history.json.')
        return

    with open(TREND_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)
    print(f'\nWrote {len(history["runs"])} run(s) to {TREND_FILE}')
    print('Next step: upload trend_history.json to the PLAYER Jira Data Quality Drive folder.')


if __name__ == '__main__':
    main()
