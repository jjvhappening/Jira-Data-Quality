"""
build_reports.py — PLAYER Data Quality Report Builder
Reads audit_results.json and produces:
  - 5 tribe-level spreadsheets  (PLAYER_DataQuality_<Tribe>_<date>.xlsx)
  - 1 master summary spreadsheet (PLAYER_DataQuality_Master_<date>.xlsx)

Run after process_full_audit.py has written audit_results.json.
"""
import json, sys, datetime, os
sys.stdout.reconfigure(encoding='utf-8')

# ── Run log ───────────────────────────────────────────────────────────────────
RUN_LOG_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'run_log.json')
TREND_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'trend_history.json')

def load_run_log():
    if os.path.exists(RUN_LOG_FILE):
        with open(RUN_LOG_FILE, encoding='utf-8') as f:
            return json.load(f)
    return []

def append_run_log(entry):
    log = load_run_log()
    log.append(entry)
    with open(RUN_LOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(log, f, indent=2, ensure_ascii=False)

# ── Trend history ──────────────────────────────────────────────────────────────
def load_trend_history():
    if os.path.exists(TREND_FILE):
        with open(TREND_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {'schema_version': 1, 'runs': []}

def save_trend_history(history):
    with open(TREND_FILE, 'w', encoding='utf-8') as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

def compute_run_entry(rows, date_str, run_num):
    total = len(rows)
    if not total:
        return None
    with_gaps  = sum(1 for r in rows if r['score'] < 100)
    compliant  = total - with_gaps
    avg_score  = round(sum(r['score'] for r in rows) / total)
    gap_counts = {}
    for lbl, fld in FIELD_LABELS:
        n = sum(1 for r in rows if fld in r.get('fields', {}) and not r['fields'][fld])
        if n:
            gap_counts[lbl] = n

    tribes_out = {}
    for tribe in TRIBES_ORDER:
        t_rows = [r for r in rows if r.get('tribe', 'Unassigned') == tribe]
        if not t_rows:
            continue
        t_total     = len(t_rows)
        t_with_gaps = sum(1 for r in t_rows if r['score'] < 100)
        t_avg       = round(sum(r['score'] for r in t_rows) / t_total)
        t_gap_c     = {}
        for lbl, fld in FIELD_LABELS:
            n = sum(1 for r in t_rows if fld in r.get('fields', {}) and not r['fields'][fld])
            if n:
                t_gap_c[lbl] = n

        squads_out = {}
        squad_map  = defaultdict(list)
        for r in t_rows:
            squad_map[normalise_squad(r.get('squad', 'Unassigned'))].append(r)
        for squad, s_rows in squad_map.items():
            s_avg   = round(sum(r['score'] for r in s_rows) / len(s_rows))
            s_with  = sum(1 for r in s_rows if r['score'] < 100)
            s_gap_c = {}
            for lbl, fld in FIELD_LABELS:
                n = sum(1 for r in s_rows if fld in r.get('fields', {}) and not r['fields'][fld])
                if n:
                    s_gap_c[lbl] = n
            # Top 5 non-compliant for this squad (powers the dashboard squad filter)
            s_active = sorted(
                [r for r in s_rows if r.get('missing')],
                key=lambda r: len(r.get('missing', [])), reverse=True
            )[:5]
            squads_out[squad] = {
                'total': len(s_rows), 'avg_score': s_avg,
                'with_gaps': s_with, 'gap_counts': s_gap_c,
                'top_offenders': [
                    {'key': r['key'], 'summary': r['summary'],
                     'score': r['score'], 'missing': r['missing']}
                    for r in s_active
                ],
            }

        tribes_out[tribe] = {
            'total': t_total, 'avg_score': t_avg,
            'with_gaps': t_with_gaps, 'compliant': t_total - t_with_gaps,
            'gap_counts': t_gap_c, 'squads': squads_out,
        }

    active = sorted(
        [r for r in rows if r.get('status', '').lower() != 'done' and r.get('missing')],
        key=lambda r: len(r.get('missing', [])), reverse=True
    )
    top_offenders = [
        {'key': r['key'], 'summary': r['summary'],
         'tribe': r.get('tribe', 'Unassigned'), 'squad': r.get('squad', ''),
         'score': r['score'], 'missing_count': len(r['missing']), 'missing': r['missing']}
        for r in active[:20]
    ]

    return {
        'run': run_num, 'date': date_str,
        'total': total, 'with_gaps': with_gaps, 'compliant': compliant,
        'avg_score': avg_score, 'gap_counts': gap_counts,
        'tribes': tribes_out, 'top_offenders': top_offenders,
        'source': 'build_reports',
        # ISO timestamp of when Jira data was fetched (from process_full_audit.py).
        # Falls back to RUN_TS (build time) if the field is absent (older audit files).
        'data_fetched_at': rows[0].get('fetched_at', RUN_TS) if rows else RUN_TS,
    }

run_log    = load_run_log()
RUN_NUMBER = len(run_log) + 1
RUN_TS     = datetime.datetime.now().strftime('%Y-%m-%dT%H:%M:%S')
print(f'Run #{RUN_NUMBER} starting at {RUN_TS}')
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
AUDIT_FILE = os.path.join(_SCRIPT_DIR, 'audit_results.json')
BASE_DIR   = _SCRIPT_DIR

# ── Squad name normalisation (mirrors build_tribe_messages.py / build_squad_messages.py)
SQUAD_NORM = {
    'Thundercats (Promo Journeys)': 'Thundercats',
    'Thundercats (Engagement Tribe)': 'Thundercats',
    'GraySkull (Bonus Platform)': 'GraySkull',
    'Retail Sports Experience (RISE)': 'Retail Sports Experience',
}

def normalise_squad(name):
    return SQUAD_NORM.get(name, name)

# ── Constants (single source of truth) ────────────────────────────────────────
FIELD_LABELS = [
    # ('Summary', 'summary') removed — Jira enforces this field; always populated.
    ('Parent',                'parent'),
    ('Product Lead',          'customfield_12121'),
    ('Engineering Lead',      'customfield_12122'),
    ('Roadmap Priority',      'customfield_12012'),
    ('Impact (PDT)',          'customfield_12178'),
    ('Country (PDT)',         'customfield_12112'),
    ('Areas Impacted',        'customfield_12110'),
    ('Investment Category',   'customfield_10709'),
    ('Start Date',            'customfield_10025'),
    ('Description / PRD',     'desc_or_prd'),
    ('Tribes Impacted',       'customfield_12109'),
    ('End of Definition Date','customfield_15460'),
    ('Health Status',         'customfield_12111'),
    ('Planned Release Date',  'customfield_12114'),
    ('Short Status Update',   'customfield_14447'),
    ('Due Date',              'duedate'),
    ('Actual End Date',       'customfield_12180'),
]

TIER_FOR_FIELD = {
    # 'summary': 0  removed — see FIELD_LABELS note above.
    'parent': 0, 'customfield_12121': 0, 'customfield_12122': 0,
    'customfield_12012': 0, 'customfield_12178': 0, 'customfield_12112': 0,
    'customfield_12110': 0, 'customfield_10709': 0,
    'customfield_10025': 1,
    'desc_or_prd': 2,
    'customfield_12109': 3, 'customfield_15460': 3,
    'customfield_12111': 4, 'customfield_12114': 4, 'customfield_14447': 4, 'duedate': 4,
    'customfield_12180': 5,
}

TRIBES_ORDER = [
    'Player Engagement', 'Transact', 'Fraud Prevention',
    'Retail/Multichannel', 'Manage', 'Unassigned',
]

TRIBE_FILENAMES = {
    'Player Engagement':  'PlayerEngagement',
    'Transact':           'Transact',
    'Fraud Prevention':   'FraudPrevention',
    'Retail/Multichannel':'RetailMultichannel',
    'Manage':             'Manage',
}

# ── Styles (defined once, shared across all sheets) ───────────────────────────
HDR_FILL  = PatternFill('solid', fgColor='1F4E79')
HDR_FONT  = Font(color='FFFFFF', bold=True)
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
PASS_FONT = Font(color='276221')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
FAIL_FONT = Font(color='9C0006')
WARN_FILL = PatternFill('solid', fgColor='FFEB9C')
NA_FILL   = PatternFill('solid', fgColor='F2F2F2')
NA_FONT   = Font(color='AAAAAA')
LINK_FONT = Font(color='0563C1', underline='single')
_thin     = Side(style='thin', color='CCCCCC')
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def score_fill(score):
    if score == 100: return PASS_FILL
    if score < 70:   return FAIL_FILL
    return WARN_FILL

def hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
    return c

def data(ws, row, col, value, align='center', fill=None, font=None, wrap=False, hyperlink=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = BORDER
    if fill:          c.fill = fill
    if font:          c.font = font
    if hyperlink:     c.hyperlink = hyperlink; c.font = LINK_FONT
    if number_format: c.number_format = number_format
    return c

def squad_jql(squad_name):
    # NOTE — Jira Cloud's issue navigator (/issues/?jql=...) does NOT support
    # pre-selecting columns via URL parameters.  There is no &columnNames= or
    # equivalent parameter that the cloud issue navigator honours; column
    # configuration is stored server-side per user and cannot be overridden
    # from a URL.  The recommended workaround is to tell recipients in the
    # accompanying Slack message which columns to add manually.  See Step 4
    # of player-jira-data-quality-workflow.md for the suggested-columns text
    # that is included in every squad Slack message.
    enc = squad_name.replace(' ', '%20').replace('/', '%2F').replace('(', '%28').replace(')', '%29')
    return (f'https://axilis.atlassian.net/issues/?jql=project%20%3D%20PLAYER'
            f'%20AND%20issuetype%20%3D%20Initiative'
            f'%20AND%20cf%5B11250%5D%20%3D%20%22{enc}%22'
            f'%20AND%20status%20NOT%20IN%20(Backlog)')

def gap_counts_for(rows):
    """Return {label: count} for fields that are required-but-missing in these rows."""
    counts = {}
    for lbl, fld in FIELD_LABELS:
        n = sum(1 for r in rows if fld in r['fields'] and not r['fields'][fld])
        counts[lbl] = n
    return counts

# ── Load data (once) ──────────────────────────────────────────────────────────
with open(AUDIT_FILE, encoding='utf-8') as f:
    audit = json.load(f)

all_rows = audit['rows']
today    = datetime.date.today().strftime('%Y-%m-%d')

OUT_DIR = os.path.join(BASE_DIR, f'PLAYER Data Quality — Run {RUN_NUMBER} — {today}')
os.makedirs(OUT_DIR, exist_ok=True)
print(f'Output folder: {OUT_DIR}')

# ═══════════════════════════════════════════════════════════════════════════════
# TRIBE SPREADSHEETS
# ═══════════════════════════════════════════════════════════════════════════════

def build_tribe_workbook(tribe_rows):
    tribe_rows = sorted(tribe_rows, key=lambda r: r['score'])
    wb = Workbook()

    # ── Detail tab ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Detail'
    headers = ['Compliance Score', 'Issue Key', 'Initiative Name', 'Status', 'Squad', 'Created'] \
              + [lbl for lbl, _ in FIELD_LABELS]
    for col, h in enumerate(headers, 1):
        hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 50
    ws.freeze_panes = 'A2'

    for ri, r in enumerate(tribe_rows, 2):
        tiers = r['tiers']
        data(ws, ri, 1, r['score'], fill=score_fill(r['score']), number_format='0"%"')
        data(ws, ri, 2, r['key'],
             hyperlink=f"https://axilis.atlassian.net/browse/{r['key']}")
        data(ws, ri, 3, r['summary'], align='left', wrap=True)
        data(ws, ri, 4, r['status'])
        data(ws, ri, 5, r['squad'])
        data(ws, ri, 6, r.get('created') or '')

        for ci, (lbl, fld) in enumerate(FIELD_LABELS):
            col = 7 + ci
            if TIER_FOR_FIELD.get(fld) not in tiers:
                data(ws, ri, col, 'N/A', fill=NA_FILL, font=NA_FONT)
            else:
                ok = r['fields'].get(fld, False)
                data(ws, ri, col, '✓' if ok else '✗',
                     fill=PASS_FILL if ok else FAIL_FILL,
                     font=PASS_FONT if ok else FAIL_FONT)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    for i in range(len(FIELD_LABELS)):
        ws.column_dimensions[get_column_letter(7+i)].width = 15

    # ── Summary tab ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    s_headers = (['Squad', 'Total', 'Avg Score']
                 + [lbl for lbl, _ in FIELD_LABELS]
                 + ['View in Jira'])
    for col, h in enumerate(s_headers, 1):
        hdr(ws2, 1, col, h)
    ws2.row_dimensions[1].height = 50
    ws2.freeze_panes = 'A2'

    squads = defaultdict(list)
    for r in tribe_rows:
        squads[r['squad']].append(r)
    squad_list = sorted(squads.items(), key=lambda x: sum(r['score'] for r in x[1])/len(x[1]))

    for ri, (squad, rows) in enumerate(squad_list, 2):
        avg = round(sum(r['score'] for r in rows) / len(rows))
        data(ws2, ri, 1, squad, align='left')
        data(ws2, ri, 2, len(rows))
        data(ws2, ri, 3, avg, fill=score_fill(avg), number_format='0"%"')
        gaps = gap_counts_for(rows)
        for ci, (lbl, fld) in enumerate(FIELD_LABELS):
            col = 4 + ci
            n = gaps.get(lbl, 0)
            cell = data(ws2, ri, col, n if n > 0 else '')
            if n > 0:
                cell.fill = FAIL_FILL; cell.font = FAIL_FONT
        data(ws2, ri, 4 + len(FIELD_LABELS), 'View in Jira',
             hyperlink=squad_jql(squad))

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 12
    for i in range(len(FIELD_LABELS)):
        ws2.column_dimensions[get_column_letter(4+i)].width = 14

    return wb


print('Building tribe spreadsheets...')
for tribe, slug in TRIBE_FILENAMES.items():
    rows = [r for r in all_rows if r.get('tribe') == tribe]
    if not rows:
        print(f'  WARN: no rows for {tribe}')
        continue
    wb = build_tribe_workbook(rows)
    path = os.path.join(OUT_DIR, f'PLAYER_DataQuality_{slug}_{today}.xlsx')
    wb.save(path)
    squads = defaultdict(list)
    for r in rows: squads[r['squad']].append(r)
    avg = round(sum(r['score'] for r in rows) / len(rows))
    print(f'  {tribe}: {len(rows)} initiatives across {len(squads)} squads, avg {avg}% → {os.path.basename(path)}')


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER SPREADSHEET
# ═══════════════════════════════════════════════════════════════════════════════

print('\nBuilding master spreadsheet...')
wb_m = Workbook()

# ── Overview tab (one row per tribe) ────────────────────────────────────────
ws_ov = wb_m.active
ws_ov.title = 'Overview'
ov_hdrs = (['Tribe', 'Total Initiatives', 'With Gaps', 'Avg Score', '% Fully Complete']
           + [lbl for lbl, _ in FIELD_LABELS])
for col, h in enumerate(ov_hdrs, 1):
    hdr(ws_ov, 1, col, h)
ws_ov.row_dimensions[1].height = 50
ws_ov.freeze_panes = 'A2'

tribe_stats = []
for tribe in TRIBES_ORDER:
    rows = [r for r in all_rows if r.get('tribe', 'Unassigned') == tribe]
    if not rows: continue
    avg = round(sum(r['score'] for r in rows) / len(rows))
    with_gaps = sum(1 for r in rows if r['score'] < 100)
    pct_full  = round(sum(1 for r in rows if r['score'] == 100) / len(rows) * 100)
    tribe_stats.append({'tribe': tribe, 'total': len(rows), 'with_gaps': with_gaps,
                        'avg': avg, 'pct_full': pct_full, 'gaps': gap_counts_for(rows)})

tribe_stats.sort(key=lambda x: x['avg'])

for ri, ts in enumerate(tribe_stats, 2):
    data(ws_ov, ri, 1, ts['tribe'], align='left')
    data(ws_ov, ri, 2, ts['total'])
    data(ws_ov, ri, 3, ts['with_gaps'])
    data(ws_ov, ri, 4, ts['avg'], fill=score_fill(ts['avg']), number_format='0"%"')
    data(ws_ov, ri, 5, ts['pct_full'], number_format='0"%"')
    for ci, (lbl, fld) in enumerate(FIELD_LABELS):
        n = ts['gaps'].get(lbl, 0)
        cell = data(ws_ov, ri, 6+ci, n if n > 0 else '')
        if n > 0: cell.fill = FAIL_FILL; cell.font = FAIL_FONT

ws_ov.column_dimensions['A'].width = 22
ws_ov.column_dimensions['B'].width = 16
ws_ov.column_dimensions['C'].width = 12
ws_ov.column_dimensions['D'].width = 12
ws_ov.column_dimensions['E'].width = 16
for i in range(len(FIELD_LABELS)):
    ws_ov.column_dimensions[get_column_letter(6+i)].width = 14

# ── By Squad tab (one row per squad across all tribes) ──────────────────────
ws_sq = wb_m.create_sheet('By Squad')
sq_hdrs = (['Squad', 'Tribe', 'Total', 'With Gaps', 'Avg Score']
           + [lbl for lbl, _ in FIELD_LABELS]
           + ['View in Jira'])
for col, h in enumerate(sq_hdrs, 1):
    hdr(ws_sq, 1, col, h)
ws_sq.row_dimensions[1].height = 50
ws_sq.freeze_panes = 'A2'

all_squads = defaultdict(list)
for r in all_rows:
    all_squads[(r['squad'], r.get('tribe', 'Unassigned'))].append(r)

sq_list = []
for (squad, tribe), rows in all_squads.items():
    avg = round(sum(r['score'] for r in rows) / len(rows))
    sq_list.append({'squad': squad, 'tribe': tribe, 'total': len(rows),
                    'with_gaps': sum(1 for r in rows if r['score'] < 100),
                    'avg': avg, 'gaps': gap_counts_for(rows)})
sq_list.sort(key=lambda x: x['avg'])

for ri, sr in enumerate(sq_list, 2):
    data(ws_sq, ri, 1, sr['squad'], align='left')
    data(ws_sq, ri, 2, sr['tribe'], align='left')
    data(ws_sq, ri, 3, sr['total'])
    data(ws_sq, ri, 4, sr['with_gaps'])
    data(ws_sq, ri, 5, sr['avg'], fill=score_fill(sr['avg']), number_format='0"%"')
    for ci, (lbl, fld) in enumerate(FIELD_LABELS):
        n = sr['gaps'].get(lbl, 0)
        cell = data(ws_sq, ri, 6+ci, n if n > 0 else '')
        if n > 0: cell.fill = FAIL_FILL; cell.font = FAIL_FONT
    data(ws_sq, ri, 6+len(FIELD_LABELS), 'View in Jira',
         hyperlink=squad_jql(sr['squad']))

ws_sq.column_dimensions['A'].width = 30
ws_sq.column_dimensions['B'].width = 20
ws_sq.column_dimensions['C'].width = 8
ws_sq.column_dimensions['D'].width = 10
ws_sq.column_dimensions['E'].width = 12
for i in range(len(FIELD_LABELS)):
    ws_sq.column_dimensions[get_column_letter(6+i)].width = 14

# ── By Field tab (gap frequency across all PLAYER) ──────────────────────────
ws_fl = wb_m.create_sheet('By Field')
fl_hdrs = ['Field', 'Initiatives Missing', '% of All Audited', 'Most Affected Tribe']
for col, h in enumerate(fl_hdrs, 1):
    hdr(ws_fl, 1, col, h)
ws_fl.row_dimensions[1].height = 40
ws_fl.freeze_panes = 'A2'

total_audited = len(all_rows)
field_stats = []
for lbl, fld in FIELD_LABELS:
    missing = [r for r in all_rows if fld in r['fields'] and not r['fields'][fld]]
    if not missing: continue
    tribe_counts = defaultdict(int)
    for r in missing:
        tribe_counts[r.get('tribe', 'Unassigned')] += 1
    top_tribe = max(tribe_counts.items(), key=lambda x: x[1])[0]
    pct = round(len(missing) / total_audited * 100)
    field_stats.append({'label': lbl, 'count': len(missing), 'pct': pct, 'top_tribe': top_tribe})

field_stats.sort(key=lambda x: -x['count'])

for ri, fs in enumerate(field_stats, 2):
    data(ws_fl, ri, 1, fs['label'], align='left')
    data(ws_fl, ri, 2, fs['count'])
    pct_fill = FAIL_FILL if fs['pct'] >= 30 else WARN_FILL if fs['pct'] >= 10 else PASS_FILL
    data(ws_fl, ri, 3, fs['pct'], fill=pct_fill, number_format='0"%"')
    data(ws_fl, ri, 4, fs['top_tribe'], align='left')

ws_fl.column_dimensions['A'].width = 24
ws_fl.column_dimensions['B'].width = 20
ws_fl.column_dimensions['C'].width = 16
ws_fl.column_dimensions['D'].width = 22

# ── Save master ─────────────────────────────────────────────────────────────
master_path = os.path.join(OUT_DIR, f'PLAYER_DataQuality_Master_{today}.xlsx')
wb_m.save(master_path)
print(f'  Saved: {os.path.basename(master_path)}')

# ── Summary ──────────────────────────────────────────────────────────────────
print(f'\n{"─"*60}')
print(f'Total audited: {total_audited} initiatives')
print(f'Top 5 gaps:')
for fs in field_stats[:5]:
    print(f'  {fs["label"]}: {fs["count"]} ({fs["pct"]}%)')
print(f'{"─"*60}')

# ── Persist run log entry ──────────────────────────────────────────────────────
tribe_counts_log = {}
for tribe in TRIBES_ORDER:
    tribe_counts_log[tribe] = len([r for r in all_rows if r.get('tribe') == tribe])

files_generated = [f'PLAYER_DataQuality_{slug}_{today}.xlsx' for slug in TRIBE_FILENAMES.values()]
files_generated.append(f'PLAYER_DataQuality_Master_{today}.xlsx')

append_run_log({
    'run': RUN_NUMBER,
    'date': today,
    'timestamp': RUN_TS,
    'total_audited': total_audited,
    'tribes': tribe_counts_log,
    'files_generated': files_generated,
    'local_folder': OUT_DIR,
    'drive_folder_id': None,
})
print(f'Run #{RUN_NUMBER} logged to run_log.json')

# ── Persist trend history entry ───────────────────────────────────────────────
entry = compute_run_entry(all_rows, today, RUN_NUMBER)
if entry:
    hist = load_trend_history()
    hist['runs'] = [r for r in hist['runs'] if r['date'] != today]
    hist['runs'].append(entry)
    hist['runs'].sort(key=lambda r: r['date'])
    save_trend_history(hist)
    print(f'Trend history updated → {os.path.basename(TREND_FILE)} ({len(hist["runs"])} runs)')

# ── Score consistency assertion ───────────────────────────────────────────────
# Verifies that each row's score is mathematically consistent with its missing
# list. Surfaces any future divergence immediately at run time so it is never
# silently propagated to the dashboard.
def assert_score_consistency(rows):
    for r in rows:
        fields = r.get('fields', {})
        if not fields:
            continue
        n_total   = len(fields)
        n_missing = sum(1 for v in fields.values() if not v)
        expected  = round((n_total - n_missing) / n_total * 100)
        if abs(expected - r['score']) > 2:
            print(f'  ⚠️  SCORE INCONSISTENCY: {r["key"]} '
                  f'score={r["score"]} but {n_missing}/{n_total} missing → expected ~{expected}')

print('Checking score consistency...')
assert_score_consistency(all_rows)
inconsistencies = [
    r for r in all_rows
    if r.get('fields') and abs(
        round((len(r['fields']) - sum(1 for v in r['fields'].values() if not v)) / len(r['fields']) * 100)
        - r['score']) > 2
]
if not inconsistencies:
    print('  ✓ All scores consistent')
else:
    print(f'  ⚠️  {len(inconsistencies)} inconsistent rows (see above)')

# ── Sync trend_history.json to Drive ─────────────────────────────────────────
# Python is the single source of truth for trend_history.json. This replaces the
# Apps Script Excel-processing pipeline that previously caused score/missing mismatches.
# drive_sync.py uses the existing clasp OAuth2 credentials (~/.clasprc.json).
# Test credentials first: python drive_sync.py --test
DRIVE_TREND_FILE_ID = '1hlDO1oo2YOT6BenZa9vqbWxBT0bRW2PA'
try:
    from drive_sync import upload_trend_history
    upload_trend_history()
except Exception as e:
    print(f'\n  ⚠️  Drive sync failed — trend_history.json NOT updated online.')
    print(f'      Error: {e}')
    print(f'      Manual fallback: upload {TREND_FILE}')
    print(f'      to Drive file ID: {DRIVE_TREND_FILE_ID}')
    print(f'      (replace existing file — do not rename it)')

print('Done.')
